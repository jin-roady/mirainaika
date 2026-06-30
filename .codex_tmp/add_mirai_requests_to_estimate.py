from copy import copy
from pathlib import Path

from openpyxl import load_workbook

path = Path("/Users/jin/Roady.net Dropbox/矢吹仁/_jp-roady/01_業務/000_みらい内科クリニック/みらい内科クリニック 美容サイト改修 見積切り分け表.xlsx")

wb = load_workbook(path)

rows = [
    [
        "2026-06-27",
        "美容修正P1",
        "お悩み説明文削除、LHALAドクター敏感肌削除、各価格表修正・並び替え",
        "未対応",
        "対象外",
        2.5,
        "既存TOP・MENU・価格表の文言削除、表記整理、並び替えが中心。",
        "美容修正P1",
    ],
    [
        "2026-06-27",
        "美容修正P2",
        "価格表の表示ずれ・回数削除、各施術の説明・同意書書体統一、FAQ削除",
        "未対応",
        "対象外",
        1.5,
        "既存価格表の表示調整、既存施術ページの書体統一・FAQ削除。",
        "美容修正P2",
    ],
    [
        "2026-06-28",
        "美容修正P3",
        "美容問診票PDF差し替え、マッサージピール・リバースピール同意書追加、PC/スマホ表示確認",
        "未対応",
        "対象外",
        2,
        "添付PDF差し替え、同意書配置、既存ページのレスポンシブ表示確認・修正。",
        "美容修正P3",
    ],
    [
        "2026-06-28",
        "ダイエット修正",
        "内服治療ダイエットのマンジャロ削除、ゼップバウンド料金・注意書き・News追加",
        "未対応",
        "対象外",
        2,
        "既存の内服治療ページの薬剤・料金・注意書き更新。現サイトと改修サイト両方が対象。",
        "ダイエット 修正",
    ],
]

mail_rows = [
    ["2026-06-27", "美容修正P1", "お悩み説明文削除、MENU・価格表修正", "メールを開く", "https://mail.google.com/mail/#all/19f049c7349d6c21"],
    ["2026-06-27", "美容修正P2", "価格表表示調整、書体統一、FAQ削除", "メールを開く", "https://mail.google.com/mail/#all/19f04a7bf04cfa59"],
    ["2026-06-28", "美容修正P3", "美容問診票PDF、同意書追加、表示確認", "メールを開く", "https://mail.google.com/mail/#all/19f0dcdacc1db97e"],
    ["2026-06-28", "ダイエット 修正", "マンジャロ削除、ゼップバウンドへ変更", "メールを開く", "https://mail.google.com/mail/#all/19f0e5d28d58ae41"],
]


def copy_row_style(ws, source_row, target_row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)
        if src.protection:
            dst.protection = copy(src.protection)


def resize_first_table(ws, ref):
    if ws.tables:
        first = next(iter(ws.tables.values()))
        first.ref = ref


outside = wb["見積対象外"]
insert_at = 16
outside.insert_rows(insert_at, len(rows))
for offset, row in enumerate(rows):
    target = insert_at + offset
    copy_row_style(outside, 15, target, 1, 8)
    for col, value in enumerate(row, start=1):
        outside.cell(target, col).value = value

subtotal_row = insert_at + len(rows)
outside.cell(subtotal_row, 6).value = 23.5
resize_first_table(outside, f"A1:H{subtotal_row}")

mail = wb["メール一覧"]
start = mail.max_row + 1
for offset, row in enumerate(mail_rows):
    target = start + offset
    copy_row_style(mail, mail.max_row if target == start else target - 1, target, 1, 4)
    for col, value in enumerate(row[:4], start=1):
        mail.cell(target, col).value = value
    mail.cell(target, 4).hyperlink = row[4]
    mail.cell(target, 4).style = "Hyperlink"

resize_first_table(mail, f"A1:D{mail.max_row}")

summary = wb["概要"]
summary["B6"] = "20件（Notion小計値 35.5〜55.5h／再計算 27.5〜40.5h）"
summary["B14"] = "Notionの小計行は再分類後に再計算されていないため、本表では各シートに「（参考）再計算合計」を併記しています。追加依頼（美容修正P1〜P3、ダイエット修正）を反映済み。"

wb.save(path)
